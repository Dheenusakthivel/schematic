import streamlit as st
import pandas as pd
import fitz  # PyMuPDF
import re
import pdfplumber
import os
import tempfile
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
import pytesseract
from pdf2image import convert_from_path
import base64
import uuid

# Streamlit page configuration
st.set_page_config(page_title="Component Analyzer", layout="wide")

# Title and description
st.title("Component Analyzer")
st.markdown("""
Upload an Excel file and a PDF schematic to analyze components. The app will:
- Merge and expand component lists in the Excel file.
- Extract components from the PDF.
- Highlight components in the PDF based on conditions.
- Generate a detailed report comparing components.
""")

# File upload section
st.header("Upload Input Files")
col1, col2 = st.columns(2)
with col1:
    excel_file = st.file_uploader("Upload Excel File (.xlsx)", type=["xlsx"])
with col2:
    pdf_file = st.file_uploader("Upload PDF Schematic (.pdf)", type=["pdf"])

# Function to validate file uploads
def validate_file(file, expected_ext):
    if file is None:
        st.error(f"Please upload a {expected_ext} file.")
        return False
    if not file.name.lower().endswith(expected_ext):
        st.error(f"File '{file.name}' must have {expected_ext} extension.")
        return False
    return True

# Proceed only if both files are uploaded
if excel_file and pdf_file:
    if not (validate_file(excel_file, '.xlsx') and validate_file(pdf_file, '.pdf')):
        st.stop()

    # Create temporary files for processing
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_excel:
        tmp_excel.write(excel_file.read())
        input_excel_file = tmp_excel.name
    with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as tmp_pdf:
        tmp_pdf.write(pdf_file.read())
        input_pdf_file = tmp_pdf.name

    # Generate output file paths
    base_name = os.path.splitext(excel_file.name)[0]
    output_excel_file = f"{base_name}_modified.xlsx"
    output_pdf_file = f"{base_name}_combined.pdf"
    output_report_file = f"{base_name}_detailed_report.xlsx"

    # Read Excel file
    try:
        df = pd.read_excel(input_excel_file, engine='openpyxl')
    except Exception as e:
        st.error(f"Error reading Excel file: {e}")
        st.stop()

    # Get component columns
    st.header("Select Component Columns")
    st.write("Select the columns containing component lists:")
    component_columns = st.multiselect(
        "Component Columns",
        options=list(df.columns),
        default=[],
        help="Choose columns that contain component lists (e.g., comma-separated component IDs)."
    )

    # Get component prefixes
    st.header("Specify Component Prefixes")
    default_prefixes = ['C', 'R', 'D', 'Q', 'U', 'L', 'Z', 'FB', 'SWITCH', 'MOV', 'LED', 'TVS', 'ESD', 'XT', 'OPD', 'OPQ', 'LCD', 'CN']
    prefixes_input = st.text_input(
        "Component Prefixes (comma-separated)",
        value=','.join(default_prefixes),
        help="Enter valid component prefixes (e.g., C,R,D). Leave empty to use defaults."
    )
    if prefixes_input:
        valid_prefixes = [prefix.strip().upper() for prefix in prefixes_input.split(',') if prefix.strip()]
        if not valid_prefixes:
            st.warning(f"No valid prefixes provided. Using default prefixes: {','.join(default_prefixes)}")
            valid_prefixes = default_prefixes
    else:
        valid_prefixes = default_prefixes

    single_letter_prefixes = [p for p in valid_prefixes if len(p) == 1]
    multi_letter_prefixes = [p for p in valid_prefixes if len(p) > 1]

    # Process button
    if st.button("Process Files"):
        with st.spinner("Processing files..."):
            # Process Excel data
            expanded_rows = []
            for _, row in df.iterrows():
                components = set()
                for col in component_columns:
                    if pd.notna(row[col]) and row[col].strip():
                        comps = [comp.strip().upper() for comp in str(row[col]).split(',') if comp.strip()]
                        components.update(comps)
                
                if components:
                    for component in sorted(components):
                        new_row = row.copy()
                        new_row['Merged Components'] = component
                        expanded_rows.append(new_row)
                else:
                    new_row = row.copy()
                    new_row['Merged Components'] = ''
                    expanded_rows.append(new_row)

            # Create new DataFrame
            new_df = pd.DataFrame(expanded_rows)
            columns = list(df.columns) + ['Merged Components']
            new_df = new_df[columns]

            # Save modified Excel
            try:
                new_df.to_excel(output_excel_file, index=False, engine='openpyxl')
            except Exception as e:
                st.error(f"Error saving modified Excel file: {e}")
                st.stop()

            # Read modified Excel
            try:
                df = pd.read_excel(output_excel_file, engine='openpyxl')
            except Exception as e:
                st.error(f"Error reading modified Excel file: {e}")
                st.stop()

            # Extract components from Excel
            excel_components = set()
            excel_counts = Counter()
            for item in df['Merged Components'].dropna().astype(str):
                components = [comp.strip().upper() for comp in item.split(",") if comp.strip()]
                excel_components.update(components)
                excel_counts.update(components)

            # Extract components from PDF
            def extract_components_from_pdf(doc, valid_prefixes):
                components = []
                for page_num, page in enumerate(doc, 1):
                    words = page.get_text("words")
                    for word_tuple in words:
                        word = re.sub(r'[^A-Za-z0-9]', '', word_tuple[4].strip()).upper()
                        if any(word.startswith(prefix) for prefix in valid_prefixes):
                            if any(re.match(r'^' + prefix + r'[0-9]+$', word) for prefix in single_letter_prefixes):
                                components.append((word, page_num, word_tuple))
                            elif any(re.match(r'^' + prefix + r'[0-9]+$', word) for prefix in multi_letter_prefixes):
                                components.append((word, page_num, word_tuple))
                
                # Try OCR if no components found
                if not components:
                    try:
                        images = convert_from_path(doc.name)
                        for page_num, image in enumerate(images, 1):
                            text = pytesseract.image_to_string(image).upper()
                            for word in re.findall(r'(?:[A-Z][0-9]+|(?:' + '|'.join(multi_letter_prefixes) + r')[0-9]+)', text):
                                if any(word.startswith(prefix) for prefix in valid_prefixes):
                                    components.append((word, page_num, None))
                    except Exception as e:
                        st.warning(f"OCR error: {e}")

                # Try pdfplumber if still no components
                if not components:
                    try:
                        with pdfplumber.open(doc.name) as pdf:
                            for page_num, page in enumerate(pdf.pages, 1):
                                text = page.extract_text().upper() if page.extract_text() else ""
                                for word in re.findall(r'(?:[A-Z][0-9]+|(?:' + '|'.join(multi_letter_prefixes) + r')[0-9]+)', text):
                                    if any(word.startswith(prefix) for prefix in valid_prefixes):
                                        components.append((word, page_num, None))
                    except Exception as e:
                        st.warning(f"pdfplumber error: {e}")
                
                return components

            # Open PDF
            try:
                doc = fitz.open(input_pdf_file)
            except Exception as e:
                st.error(f"Error opening PDF file: {e}")
                st.stop()

            # Extract PDF components
            pdf_components_list = extract_components_from_pdf(doc, valid_prefixes)
            pdf_components = set(comp[0] for comp in pdf_components_list)
            pdf_counts = Counter(comp[0] for comp in pdf_components_list)

            # Analyze conditions
            repeated_pdf = {comp for comp, count in pdf_counts.items() if count > 1}
            repeated_excel = {comp for comp, count in excel_counts.items() if count > 1}
            in_pdf_not_excel = pdf_components - excel_components
            in_excel_not_pdf = excel_components - pdf_components

            def filter_components(components):
                return {comp for comp in components if re.match(r'^(?:[A-Z][0-9]+|(?:' + '|'.join(valid_prefixes) + r')[0-9]+)$', comp.upper())}

            repeated_pdf = filter_components(repeated_pdf)
            repeated_excel = filter_components(repeated_excel)
            in_pdf_not_excel = filter_components(in_pdf_not_excel)
            in_excel_not_pdf = filter_components(in_excel_not_pdf)

            # Highlight components in PDF
            def highlight_components(doc, pdf_components_list, repeated_pdf, repeated_excel, in_pdf_not_excel, in_excel_not_pdf):
                for word, page_num, word_tuple in pdf_components_list:
                    if not word_tuple:
                        continue
                    page = doc[page_num - 1]
                    rect = fitz.Rect(word_tuple[0], word_tuple[1], word_tuple[2], word_tuple[3])
                    rect.x0 -= 2
                    rect.x1 += 2
                    rect.y0 -= 2
                    rect.y1 += 2
                    
                    if word in in_pdf_not_excel:
                        color = (1, 1, 0)  # Yellow
                    elif word in in_excel_not_pdf:
                        color = (0, 0, 1)  # Blue
                    elif word in repeated_pdf:
                        color = (1, 0, 0)  # Red
                    elif word in repeated_excel:
                        color = (0, 1, 1)  # Cyan
                    else:
                        color = (0, 1, 0)  # Green
                    
                    highlight = page.add_highlight_annot(rect)
                    highlight.set_colors(stroke=color)
                    highlight.set_opacity(0.5)
                    highlight.update()

            highlight_components(doc, pdf_components_list, repeated_pdf, repeated_excel, in_pdf_not_excel, in_excel_not_pdf)

            # Add summary page to PDF
            page = doc.new_page()
            text = (
                "Component Analysis Summary (All Pages)\n\n"
                f"Repeated in PDF ({len(repeated_pdf)}):\n{', '.join(sorted(repeated_pdf)) or 'None'}\n\n"
                f"Repeated in Excel ({len(repeated_excel)}):\n{', '.join(sorted(repeated_excel)) or 'None'}\n\n"
                f"Present in PDF, not in Excel ({len(in_pdf_not_excel)}):\n{', '.join(sorted(in_pdf_not_excel)) or 'None'}\n\n"
                f"Present in Excel, not in PDF ({len(in_excel_not_pdf)}):\n{', '.join(sorted(in_excel_not_pdf)) or 'None'}"
            )
            page.insert_text(
                fitz.Point(50, 50),
                text,
                fontname="helv",
                fontsize=10,
                color=(0, 0, 0)
            )

            # Save modified PDF
            try:
                doc.save(output_pdf_file)
                doc.close()
            except Exception as e:
                st.error(f"Error saving PDF: {e}")
                st.stop()

            # Generate detailed report
            wb = Workbook()
            wb.remove(wb.active)
            conditions = {
                'repeated_pdf': ('Repeated in PDF', 'Red', (1, 0, 0)),
                'repeated_excel': ('Repeated in Excel', 'Cyan', (0, 1, 1)),
                'in_pdf_not_excel': ('In PDF, not in Excel', 'Yellow', (1, 1, 0)),
                'in_excel_not_pdf': ('In Excel, not in PDF', 'Blue', (0, 0, 1)),
                'normal': ('Normal', 'Green', (0, 1, 0))
            }
            fills = {
                'repeated_pdf': PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid'),
                'repeated_excel': PatternFill(start_color='FF00FFFF', end_color='FF00FFFF', fill_type='solid'),
                'in_pdf_not_excel': PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid'),
                'in_excel_not_pdf': PatternFill(start_color='FF0000FF', end_color='FF0000FF', fill_type='solid'),
                'normal': PatternFill(start_color='FF00FF00', end_color='FF00FF00', fill_type='solid')
            }
            condition_components = {
                'repeated_pdf': repeated_pdf,
                'repeated_excel': repeated_excel,
                'in_pdf_not_excel': in_pdf_not_excel,
                'in_excel_not_pdf': in_excel_not_pdf,
                'normal': (pdf_components | excel_components) - (repeated_pdf | repeated_excel | in_pdf_not_excel | in_excel_not_pdf)
            }

            for condition_key, (condition_name, color_name, _) in conditions.items():
                ws = wb.create_sheet(title=condition_name)
                headers = ["Condition", "Component", "Number of Times Repeated", "Highlight Color"]
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col)
                    cell.value = header
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center')
                
                components = sorted(condition_components[condition_key])
                row = 2
                for comp in components:
                    repeat_count = (
                        pdf_counts.get(comp, 0) if condition_key in ['repeated_pdf', 'in_pdf_not_excel'] else
                        excel_counts.get(comp, 0) if condition_key in ['repeated_excel', 'in_excel_not_pdf'] else
                        max(pdf_counts.get(comp, 0), excel_counts.get(comp, 0))
                    )
                    ws.cell(row=row, column=1).value = condition_name
                    ws.cell(row=row, column=2).value = comp
                    ws.cell(row=row, column=3).value = repeat_count
                    ws.cell(row=row, column=4).value = color_name
                    for col in range(1, 5):
                        ws.cell(row=row, column=col).fill = fills[condition_key]
                    row += 1
                
                for col in ws.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = max_length + 2
                    ws.column_dimensions[column].width = adjusted_width

            try:
                wb.save(output_report_file)
            except Exception as e:
                st.error(f"Error generating report: {e}")
                st.stop()

            # Display summary
            st.header("Analysis Summary")
            st.write(f"**Repeated in PDF**: {len(repeated_pdf)} components")
            st.write(f"**Repeated in Excel**: {len(repeated_excel)} components")
            st.write(f"**In PDF, not in Excel**: {len(in_pdf_not_excel)} components")
            st.write(f"**In Excel, not in PDF**: {len(in_excel_not_pdf)} components")

            # Provide download links
            st.header("Download Output Files")
            for file_path, label in [
                (output_excel_file, "Modified Excel File"),
                (output_pdf_file, "Highlighted PDF"),
                (output_report_file, "Detailed Report")
            ]:
                with open(file_path, "rb") as f:
                    file_data = f.read()
                b64 = base64.b64encode(file_data).decode()
                href = f'<a href="data:application/octet-stream;base64,{b64}" download="{os.path.basename(file_path)}">Download {label}</a>'
                st.markdown(href, unsafe_allow_html=True)

            # Clean up temporary files
            os.unlink(input_excel_file)
            os.unlink(input_pdf_file)
            for file in [output_excel_file, output_pdf_file, output_report_file]:
                if os.path.exists(file):
                    os.unlink(file)

else:
    st.info("Please upload both an Excel file and a PDF file to proceed.")
    